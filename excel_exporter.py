from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER


class ExcelExporter:
    """Excel导出模块 - 使用openpyxl实现带格式的Excel导出"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def _create_styles(self):
        """创建样式定义"""
        # 表头样式
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_style = {
            'font': header_font,
            'fill': header_fill,
            'alignment': header_alignment,
            'border': header_border
        }
        
        # 数据单元格样式
        data_font = Font(size=11)
        data_alignment = Alignment(vertical='center')
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.data_style = {
            'font': data_font,
            'alignment': data_alignment,
            'border': data_border
        }
        
        # 金额列样式（右对齐）
        amount_alignment = Alignment(horizontal='right', vertical='center')
        self.amount_style = {
            'font': data_font,
            'alignment': amount_alignment,
            'border': data_border,
            'number_format': FORMAT_NUMBER
        }
        
        # 合计行样式
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.total_style = {
            'font': total_font,
            'fill': total_fill,
            'alignment': data_alignment,
            'border': data_border
        }
        
        # 合计金额样式
        self.total_amount_style = {
            'font': total_font,
            'fill': total_fill,
            'alignment': amount_alignment,
            'border': data_border,
            'number_format': FORMAT_NUMBER
        }
    
    def _apply_style(self, cell, style_dict):
        """应用样式到单元格"""
        for key, value in style_dict.items():
            if key == 'number_format':
                cell.number_format = value
            else:
                setattr(cell, key, value)
    
    def export(self, entries, file_path, include_total=True):
        """导出数据到Excel文件"""
        # 创建工作簿和工作表
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "账目数据"
        
        # 创建样式
        self._create_styles()
        
        # 定义列标题
        headers = ['日期', '人物', '描述', '金额', '类型']
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            self._apply_style(cell, self.header_style)
        
        # 写入数据行
        row_num = 2
        total_lend = 0
        total_receive = 0
        
        for entry in entries:
            # 日期
            cell = self.ws.cell(row=row_num, column=1, value=self._format_date(entry['date']))
            self._apply_style(cell, self.data_style)
            
            # 人物
            cell = self.ws.cell(row=row_num, column=2, value=entry['person'])
            self._apply_style(cell, self.data_style)
            
            # 描述
            cell = self.ws.cell(row=row_num, column=3, value=entry['description'])
            self._apply_style(cell, self.data_style)
            
            # 金额
            cell = self.ws.cell(row=row_num, column=4, value=entry['amount'])
            self._apply_style(cell, self.amount_style)
            
            # 类型
            cell = self.ws.cell(row=row_num, column=5, value=entry['type'])
            self._apply_style(cell, self.data_style)
            
            # 累计金额
            if entry['type'] == '借出':
                total_lend += entry['amount']
            else:
                total_receive += entry['amount']
            
            row_num += 1
        
        # 添加合计行
        if include_total and entries:
            # 合计标签
            cell = self.ws.cell(row=row_num, column=1, value='合计')
            self._apply_style(cell, self.total_style)
            
            # 空单元格
            for col in range(2, 4):
                cell = self.ws.cell(row=row_num, column=col)
                self._apply_style(cell, self.total_style)
            
            # 总金额
            cell = self.ws.cell(row=row_num, column=4, value=total_lend + total_receive)
            self._apply_style(cell, self.total_amount_style)
            
            # 空单元格
            cell = self.ws.cell(row=row_num, column=5)
            self._apply_style(cell, self.total_style)
            
            row_num += 1
            
            # 添加借出/收回明细
            detail_rows = [
                ('借出合计', total_lend),
                ('收回合计', total_receive),
                ('结余', total_receive - total_lend)
            ]
            
            for label, amount in detail_rows:
                cell = self.ws.cell(row=row_num, column=1, value=label)
                self._apply_style(cell, self.total_style)
                
                for col in range(2, 4):
                    cell = self.ws.cell(row=row_num, column=col)
                    self._apply_style(cell, self.total_style)
                
                cell = self.ws.cell(row=row_num, column=4, value=amount)
                self._apply_style(cell, self.total_amount_style)
                
                cell = self.ws.cell(row=row_num, column=5)
                self._apply_style(cell, self.total_style)
                
                row_num += 1
        
        # 设置列宽
        column_widths = [12, 12, 35, 12, 10]
        for col, width in enumerate(column_widths, 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width
        
        # 冻结首行
        self.ws.freeze_panes = 'A2'
        
        # 设置打印区域和标题
        self.ws.print_title_rows = '1:1'
        self.ws.page_setup.fitToWidth = True
        self.ws.page_setup.fitToHeight = False
        
        # 保存文件
        self.wb.save(file_path)
    
    def export_by_person(self, entries, file_path):
        """按人名分组导出"""
        # 按人名分组
        person_groups = {}
        for entry in entries:
            person = entry['person']
            if person not in person_groups:
                person_groups[person] = []
            person_groups[person].append(entry)
        
        # 创建工作簿
        self.wb = Workbook()
        
        # 删除默认工作表
        self.wb.remove(self.wb.active)
        
        # 创建样式
        self._create_styles()
        
        headers = ['日期', '描述', '金额', '类型']
        
        for person, person_entries in person_groups.items():
            # 创建工作表
            ws = self.wb.create_sheet(title=person[:30])  # 限制工作表名称长度
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                self._apply_style(cell, self.header_style)
            
            # 写入数据
            row_num = 2
            total_lend = 0
            total_receive = 0
            
            for entry in person_entries:
                cell = ws.cell(row=row_num, column=1, value=self._format_date(entry['date']))
                self._apply_style(cell, self.data_style)
                
                cell = ws.cell(row=row_num, column=2, value=entry['description'])
                self._apply_style(cell, self.data_style)
                
                cell = ws.cell(row=row_num, column=3, value=entry['amount'])
                self._apply_style(cell, self.amount_style)
                
                cell = ws.cell(row=row_num, column=4, value=entry['type'])
                self._apply_style(cell, self.data_style)
                
                if entry['type'] == '借出':
                    total_lend += entry['amount']
                else:
                    total_receive += entry['amount']
                
                row_num += 1
            
            # 添加合计
            cell = ws.cell(row=row_num, column=1, value='借出合计')
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=2)
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=3, value=total_lend)
            self._apply_style(cell, self.total_amount_style)
            cell = ws.cell(row=row_num, column=4)
            self._apply_style(cell, self.total_style)
            row_num += 1
            
            cell = ws.cell(row=row_num, column=1, value='收回合计')
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=2)
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=3, value=total_receive)
            self._apply_style(cell, self.total_amount_style)
            cell = ws.cell(row=row_num, column=4)
            self._apply_style(cell, self.total_style)
            row_num += 1
            
            cell = ws.cell(row=row_num, column=1, value='结余')
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=2)
            self._apply_style(cell, self.total_style)
            cell = ws.cell(row=row_num, column=3, value=total_receive - total_lend)
            self._apply_style(cell, self.total_amount_style)
            cell = ws.cell(row=row_num, column=4)
            self._apply_style(cell, self.total_style)
            
            # 设置列宽
            column_widths = [12, 35, 12, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # 冻结首行
            ws.freeze_panes = 'A2'
        
        # 保存文件
        self.wb.save(file_path)
    
    def _format_date(self, date_str):
        """格式化日期显示"""
        if len(date_str) == 8 and date_str.isdigit():
            return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        return date_str